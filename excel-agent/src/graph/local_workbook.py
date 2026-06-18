from __future__ import annotations
import logging
from pathlib import Path
from typing import Any
import openpyxl

logger = logging.getLogger(__name__)


class LocalWorkbookReader:
    """Lê planilha .xlsx local — funciona sem Azure/Graph API."""

    def __init__(self, file_path: Path, table_name: str = "", key_column: str = "") -> None:
        self._path = file_path
        self._table_name = table_name
        self._key_column = key_column

    def _find_sheet(self, wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
        if self._table_name and self._table_name in wb.sheetnames:
            return wb[self._table_name]
        # Usa a primeira aba com dados
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.max_row > 1:
                logger.info(f"Usando aba: '{name}'")
                return ws
        return wb.active

    async def get_rows(self) -> list[dict[str, Any]]:
        if not self._path.exists():
            raise FileNotFoundError(
                f"Arquivo não encontrado: {self._path}\n"
                f"Baixe a planilha do SharePoint e salve neste caminho."
            )

        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        ws = self._find_sheet(wb)

        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter, []))]

        rows = []
        for row_values in rows_iter:
            if all(v is None for v in row_values):
                continue
            record = {headers[i]: row_values[i] for i in range(min(len(headers), len(row_values)))}
            rows.append(record)

        wb.close()
        logger.info(f"Lidos {len(rows)} registros de '{self._path.name}'")
        return rows

    def detect_key_column(self, rows: list[dict[str, Any]]) -> str:
        if self._key_column:
            return self._key_column
        if not rows:
            return "__row_index__"
        for col in rows[0].keys():
            values = [r.get(col) for r in rows if r.get(col) is not None]
            if len(values) == len(rows) and len(set(str(v) for v in values)) == len(values):
                logger.info(f"Coluna-chave detectada automaticamente: '{col}'")
                return col
        return "__row_index__"
